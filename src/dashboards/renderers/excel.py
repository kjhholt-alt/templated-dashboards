"""Excel renderer.

Maps IR components to openpyxl primitives:

* Header: merged cell row with title + subtitle.
* Section: bold uppercase row with thin border bottom.
* status_card: row with coloured tier cell + project + summary + updated.
* kpi_tile: 2x2 block (label / value / delta) with tone fill on value.
* table: header row (frozen) + rows + thin bottom borders.
* timeline: rows ordered top-to-bottom with date / title / detail and
  a leading colour cell.
* callout: full-width row with tone left border.
* chart: caption row + per-series rows; openpyxl native bar/line chart
  if data is fully numeric.

Requires ``openpyxl``. Raises ``ImportError`` with a clear message if
the optional extra isn't installed.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from ..components import kpi_tile as kpi_helpers
from ..components import status_card as status_helpers
from ..components import table as table_helpers
from ..ir import (
    Callout,
    Chart,
    CodeBlock,
    Dashboard,
    KPITile,
    LinkGrid,
    Pipeline,
    Section,
    StatusCard,
    Table,
    Timeline,
)


def render_excel(dash: Dashboard, out: Path, *, sheet_name: str = "Dashboard") -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "Excel rendering requires openpyxl. Install with: "
            "pip install templated-dashboards[excel]"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    palette = _palette(dash.theme)
    thin = Side(style="thin", color=palette["line"])
    border_bottom = Border(bottom=thin)

    # Column widths (rough defaults)
    for i, w in enumerate([28, 28, 28, 28, 28, 28], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    ws.cell(row=row, column=1, value=dash.title).font = Font(
        bold=True, size=16, color=palette["fg0"]
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    if dash.subtitle:
        c = ws.cell(row=row, column=1, value=dash.subtitle)
        c.font = Font(size=10, italic=True, color=palette["fg2"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1

    for section in dash.sections:
        row = _emit_section(
            ws, row, section, palette, thin, border_bottom,
            Font, PatternFill, Alignment, Border,
        )
        row += 1

    if dash.footer:
        row += 1
        c = ws.cell(row=row, column=1, value=dash.footer)
        c.font = Font(size=9, color=palette["fg3"], italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    ws.freeze_panes = "A4"

    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _palette(theme: str) -> dict:
    if theme == "light":
        return dict(
            bg0="FFF5F7FA", bg1="FFFFFFFF", bg2="FFEEF2F7",
            fg0="FF0F172A", fg1="FF1E293B", fg2="FF475569", fg3="FF94A3B8",
            line="FFD0D7E2",
            good="FF15803D", warn="FFB45309", bad="FFB91C1C", neutral="FF64748B",
        )
    # palantir (default)
    return dict(
        bg0="FF07090C", bg1="FF0C1117", bg2="FF111722",
        fg0="FFE6EDF3", fg1="FFB3C1D1", fg2="FF6F8197", fg3="FF45596F",
        line="FF1F2A3B",
        good="FF22C55E", warn="FFF59E0B", bad="FFEF4444", neutral="FF94A3B8",
    )


def _tone_fill(palette: dict, tone: str, PatternFill):
    return PatternFill(
        fill_type="solid", start_color=palette.get(tone, palette["neutral"]),
        end_color=palette.get(tone, palette["neutral"]),
    )


def _emit_section(
    ws, row, section: Section, palette, thin, border_bottom,
    Font, PatternFill, Alignment, Border,
) -> int:
    c = ws.cell(row=row, column=1, value=section.title.upper())
    c.font = Font(bold=True, size=11, color=palette["fg2"])
    c.border = border_bottom
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    if section.subtitle:
        c = ws.cell(row=row, column=1, value=section.subtitle)
        c.font = Font(size=9, italic=True, color=palette["fg3"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1

    for comp in section.components:
        row = _emit_component(
            ws, row, comp, palette, thin, border_bottom,
            Font, PatternFill, Alignment, Border,
        )
        row += 1
    return row


def _emit_component(
    ws, row, comp, palette, thin, border_bottom,
    Font, PatternFill, Alignment, Border,
) -> int:
    if isinstance(comp, StatusCard):
        return _emit_status(ws, row, comp, palette, Font, PatternFill)
    if isinstance(comp, KPITile):
        return _emit_kpi(ws, row, comp, palette, Font, PatternFill)
    if isinstance(comp, Table):
        return _emit_table(ws, row, comp, palette, Font, PatternFill, border_bottom)
    if isinstance(comp, Timeline):
        return _emit_timeline(ws, row, comp, palette, Font, PatternFill)
    if isinstance(comp, Callout):
        return _emit_callout(ws, row, comp, palette, Font, PatternFill)
    if isinstance(comp, Chart):
        return _emit_chart(ws, row, comp, palette, Font, PatternFill)
    if isinstance(comp, Pipeline):
        return _emit_pipeline(ws, row, comp, palette, Font, PatternFill)
    if isinstance(comp, LinkGrid):
        return _emit_link_grid(ws, row, comp, palette, Font, PatternFill)
    if isinstance(comp, CodeBlock):
        return _emit_code_block(ws, row, comp, palette, Font, PatternFill)
    return row


def _emit_status(ws, row, c: StatusCard, palette, Font, PatternFill) -> int:
    r = status_helpers.resolve(c)
    tier = r["tier"]
    tone_cell = ws.cell(row=row, column=1, value="")
    tone_cell.fill = _tone_fill(palette, tier, PatternFill)
    ws.cell(row=row, column=2, value=r["project"] or c.project).font = Font(bold=True)
    ws.cell(row=row, column=3, value=r["summary"] or "")
    ws.cell(row=row, column=6, value=r["last_update"] or "").font = Font(
        italic=True, color=palette["fg3"]
    )
    return row


def _emit_kpi(ws, row, c: KPITile, palette, Font, PatternFill) -> int:
    label = ws.cell(row=row, column=1, value=c.label)
    label.font = Font(size=9, color=palette["fg2"], bold=True)
    val_cell = ws.cell(row=row + 1, column=1, value=_kpi_native(c))
    val_cell.font = Font(size=18, bold=True, color=palette.get(c.tone, palette["fg0"]))
    if c.delta:
        d = ws.cell(row=row + 2, column=1, value=c.delta)
        d.font = Font(size=9, color=palette["fg2"])
        return row + 2
    return row + 1


def _kpi_native(c: KPITile):
    if isinstance(c.value, (int, float)):
        return c.value
    # try int / float coerce
    s = str(c.value).replace(",", "").strip()
    try:
        if "." in s:
            return float(s)
        return int(s)
    except Exception:
        return c.value


def _emit_table(ws, row, c: Table, palette, Font, PatternFill, border_bottom) -> int:
    if c.caption:
        cap = ws.cell(row=row, column=1, value=c.caption)
        cap.font = Font(bold=True, italic=True, color=palette["fg2"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    header_row = row
    for i, h in enumerate(c.headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, color=palette["fg2"], size=10)
        cell.fill = PatternFill(
            fill_type="solid", start_color=palette["bg2"], end_color=palette["bg2"]
        )
        cell.border = border_bottom
    row += 1

    rows = table_helpers.normalise_rows(c)
    for r_data in rows:
        for i, val in enumerate(r_data, start=1):
            ws.cell(row=row, column=i, value=val).border = border_bottom
        row += 1

    # Freeze table header by setting freeze_panes if first table on sheet.
    if ws.freeze_panes is None:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    return row - 1


def _emit_timeline(ws, row, c: Timeline, palette, Font, PatternFill) -> int:
    last = row
    for ev in c.events:
        marker = ws.cell(row=row, column=1, value="")
        marker.fill = _tone_fill(palette, ev.tone, PatternFill)
        ws.cell(row=row, column=2, value=ev.when).font = Font(
            color=palette["fg3"], size=10
        )
        ws.cell(row=row, column=3, value=ev.title).font = Font(bold=True)
        if ev.detail:
            ws.cell(row=row, column=4, value=ev.detail)
        last = row
        row += 1
    return last


def _emit_callout(ws, row, c: Callout, palette, Font, PatternFill) -> int:
    bar = ws.cell(row=row, column=1, value="")
    bar.fill = _tone_fill(palette, c.tone, PatternFill)
    txt = ws.cell(row=row, column=2, value=c.text)
    txt.font = Font(color=palette["fg1"], italic=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    return row


def _emit_chart(ws, row, c: Chart, palette, Font, PatternFill) -> int:
    # Caption
    if c.caption:
        cap = ws.cell(row=row, column=1, value=c.caption)
        cap.font = Font(bold=True, color=palette["fg2"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    # Header: series labels along columns
    ws.cell(row=row, column=1, value="idx").font = Font(bold=True)
    for i, s in enumerate(c.series, start=2):
        ws.cell(row=row, column=i, value=s.label).font = Font(bold=True)
    header_row = row
    row += 1

    max_n = max(len(s.points) for s in c.series) if c.series else 0
    for j in range(max_n):
        ws.cell(row=row, column=1, value=j + 1)
        for i, s in enumerate(c.series, start=2):
            if j < len(s.points):
                ws.cell(row=row, column=i, value=s.points[j])
        row += 1

    # Native Excel chart
    try:
        from openpyxl.chart import BarChart, LineChart, Reference

        ChartCls = BarChart if c.kind == "bar" else LineChart
        ch = ChartCls()
        ch.title = c.caption or ""
        ch.legend = None
        data_ref = Reference(
            ws, min_col=2, max_col=1 + len(c.series),
            min_row=header_row, max_row=header_row + max_n,
        )
        cats_ref = Reference(
            ws, min_col=1, max_col=1,
            min_row=header_row + 1, max_row=header_row + max_n,
        )
        ch.add_data(data_ref, titles_from_data=True)
        ch.set_categories(cats_ref)
        ws.add_chart(ch, ws.cell(row=header_row, column=2 + len(c.series) + 1).coordinate)
    except Exception:  # pragma: no cover - graceful degradation
        pass

    return row - 1


def _emit_pipeline(ws, row, c: Pipeline, palette, Font, PatternFill) -> int:
    if c.caption:
        cap = ws.cell(row=row, column=1, value=c.caption)
        cap.font = Font(bold=True, color=palette["fg2"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    last = row
    for i, st in enumerate(c.stages, start=1):
        ws.cell(row=row, column=1, value=f"{i:02d}").font = Font(
            color=palette["fg3"], size=9
        )
        ws.cell(row=row, column=2, value=st.name).font = Font(
            bold=True, color=palette["fg1"]
        )
        ws.cell(row=row, column=3, value=str(st.value)).font = Font(
            bold=True,
            color=palette.get(_state_to_tone(st.state), palette["fg0"]),
        )
        ws.cell(row=row, column=4, value=st.state).font = Font(
            color=palette["fg2"], size=10
        )
        if st.detail:
            ws.cell(row=row, column=5, value=st.detail).font = Font(
                color=palette["fg2"], size=10
            )
        last = row
        row += 1
    return last


def _state_to_tone(state: str) -> str:
    if state == "ready":
        return "good"
    if state == "watching":
        return "warn"
    if state == "blocked":
        return "bad"
    return "neutral"


def _emit_link_grid(ws, row, c: LinkGrid, palette, Font, PatternFill) -> int:
    if c.caption:
        cap = ws.cell(row=row, column=1, value=c.caption)
        cap.font = Font(bold=True, color=palette["fg2"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    last = row
    for it in c.items:
        if it.kicker:
            ws.cell(row=row, column=1, value=it.kicker).font = Font(
                color=palette["fg3"], size=9, italic=True
            )
        cell = ws.cell(row=row, column=2, value=it.label)
        cell.font = Font(bold=True, color=palette["fg0"])
        cell.hyperlink = it.href
        cell.style = "Hyperlink"
        if it.detail:
            ws.cell(row=row, column=3, value=it.detail).font = Font(
                color=palette["fg2"], size=10
            )
        if it.ok is True:
            ws.cell(row=row, column=6, value="OK").font = Font(
                color=palette["good"], bold=True, size=9
            )
        elif it.ok is False:
            ws.cell(row=row, column=6, value="MISSING").font = Font(
                color=palette["bad"], bold=True, size=9
            )
        last = row
        row += 1
    return last


def _emit_code_block(ws, row, c: CodeBlock, palette, Font, PatternFill) -> int:
    if c.caption:
        cap = ws.cell(row=row, column=1, value=c.caption)
        cap.font = Font(bold=True, color=palette["fg2"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    cell = ws.cell(row=row, column=1, value=c.text or "")
    cell.font = Font(name="Consolas", size=9, color=palette["fg2"])
    cell.alignment = __import__(
        "openpyxl.styles", fromlist=["Alignment"]
    ).Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    return row


__all__ = ["render_excel"]
